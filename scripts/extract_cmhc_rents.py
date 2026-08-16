"""
Extract CMHC average rents by zone and bedroom type into a tidy CSV.

Table 1.1.2 of the Rental Market Survey reports average rent for each zone
across four bedroom types, for two survey years per workbook. The layout is a
two-row header: bedroom type spans several columns on one row, survey dates
sit beneath it, and every value is followed by a reliability-code column
(a/b/c/d) that has to be skipped.

Only the later year in each workbook is kept, so the newest file wins.

Run from anywhere:
    venv/bin/python scripts/extract_cmhc_rents.py
"""

import os
import re

import openpyxl
import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_FOLDER = os.path.join(PROJECT_ROOT, "data", "raw")
OUTPUT_FILE = os.path.join(
    PROJECT_ROOT, "data", "processed", "cmhc_average_rents.csv"
)

SHEET = "Table 1.1.2"

BEDROOM_LABELS = {
    "studio": 0,
    "bachelor": 0,
    "1 bedroom": 1,
    "2 bedroom": 2,
    "3 bedroom +": 3,
    "3 bedroom+": 3,
    "3 bedroom": 3,
}


def _parse_rent(value):
    """CMHC uses '**' and '-' for suppressed or unavailable figures."""
    if value is None:
        return None
    text = str(value).replace(",", "").replace("$", "").strip()
    if not re.fullmatch(r"\d+(\.\d+)?", text):
        return None
    return int(float(text))


def _column_map(header_rows):
    """
    Map each data column to (bedrooms, survey_date).

    The bedroom label appears once above its group and is blank for the rest of
    the span, so it carries forward until the next label.
    """
    bedroom_row, date_row = header_rows
    mapping = {}
    current = None

    for idx, label in enumerate(bedroom_row):
        if label:
            key = str(label).strip().lower()
            current = BEDROOM_LABELS.get(key, current)

        date = date_row[idx] if idx < len(date_row) else None
        if current is not None and date:
            date_text = str(date).strip()
            # Reliability-code columns have no date beneath them.
            if re.search(r"\d{2}", date_text):
                mapping[idx] = (current, date_text)

    return mapping


def extract_workbook(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        return []

    sheet = workbook[SHEET]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]

    header_index = next(
        i for i, row in enumerate(rows)
        if any(str(c).strip().lower() == "zone" for c in row if c)
    )
    mapping = _column_map((rows[header_index - 1], rows[header_index]))

    # Keep only the most recent survey date in this workbook.
    latest = max(date for _, date in mapping.values())

    records = []
    for row in rows[header_index + 1:]:
        zone = row[0]
        if not zone or not str(zone).strip():
            continue

        for column, (bedrooms, date) in mapping.items():
            if date != latest or column >= len(row):
                continue
            rent = _parse_rent(row[column])
            if rent is not None:
                records.append(
                    {
                        "Zone": str(zone).strip(),
                        "Bedrooms": bedrooms,
                        "Average_Rent": rent,
                        "Survey": latest,
                    }
                )

    return records


def main():
    all_records = []
    for filename in sorted(os.listdir(RAW_FOLDER)):
        if not filename.endswith(".xlsx"):
            continue
        found = extract_workbook(os.path.join(RAW_FOLDER, filename))
        print(f"{filename}: {len(found)} rent figures")
        all_records.extend(found)

    if not all_records:
        print("No rent data found.")
        return

    frame = pd.DataFrame(all_records)

    # Newest survey wins where workbooks overlap.
    frame = (
        frame.sort_values("Survey")
        .drop_duplicates(subset=["Zone", "Bedrooms"], keep="last")
        .sort_values(["Zone", "Bedrooms"])
        .reset_index(drop=True)
    )

    frame.to_csv(OUTPUT_FILE, index=False)
    print(f"\nSaved {len(frame)} rows to {OUTPUT_FILE}")
    print(frame.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
